import core.grpc_module as grpc_module
from importlib import import_module
import grpc
import core.payload as payload
import copy
import time
import asyncio
import pandas as pd
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


RED    = "\033[91m"
GREEN  = "\033[92m"
YELLOW = "\033[93m"
CYAN   = "\033[96m"
BOLD   = "\033[1m"
DIM    = "\033[2m"
RESET  = "\033[0m"
SEP    = "=" * 64


# ── Helpers ────────────────────────────────────────────────────────────────────

def _mask_metadata(metadata):
    if not metadata:
        return "None"
    _sensitive = {'authorization', 'token', 'secret', 'password', 'api-key', 'apikey'}
    parts = []
    for key, value in metadata:
        masked = (value[:6] + "***") if len(value) > 6 else "***"
        parts.append(f"{key}: {masked}" if any(s in key.lower() for s in _sensitive)
                     else f"{key}: {value}")
    return "  |  ".join(parts)


def _print_banner(url, pathname, metadata=None, filters=None, proxy=None):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    meta_line   = f"\n  {BOLD}Metadata:{RESET} {_mask_metadata(metadata)}" if metadata else ""
    proxy_line  = f"\n  {BOLD}Proxy   :{RESET} {proxy}" if proxy else ""
    filter_parts = []
    if filters:
        if filters.get('attacks'):
            filter_parts.append(f"attacks={','.join(filters['attacks'])}")
        if filters.get('services'):
            filter_parts.append(f"services={','.join(filters['services'])}")
    filter_line = f"\n  {BOLD}Filters :{RESET} {' | '.join(filter_parts)}" if filter_parts else ""
    print(f"""
{BOLD}{CYAN}{SEP}
  GSTF - gRPC Security Testing Framework
  Swiss German University | MIT Thesis 2025
{SEP}{RESET}
  {BOLD}Target  :{RESET} {url}
  {BOLD}Proto   :{RESET} {pathname}
  {BOLD}Started :{RESET} {now}{meta_line}{proxy_line}{filter_line}
""")


def _calculate_confidence(matches, is_time_based):
    if not matches and not is_time_based:
        return 0
    score = 0
    if matches:
        score = 55 + min(len(matches) - 1, 3) * 8
    if is_time_based:
        score = max(score, 65) + 10
    return min(score, 95)


def _confidence_label(score):
    if score >= 75: return "High"
    if score >= 55: return "Medium"
    return "Low"


# ── Auth Flow ──────────────────────────────────────────────────────────────────

async def _run_auth_flow(stub_map, gs_service, auth_rpc, auth_data, auth_field,
                         auth_header, timeout):
    """Fix #2: uses stub_map to find the correct stub for the auth RPC."""
    service_name = auth_rpc + "Request"
    stub = stub_map.get(service_name)
    if stub is None:
        print(f"  {RED}[!] '{service_name}' not found in proto{RESET}")
        return None

    try:
        module = getattr(gs_service, service_name)
    except AttributeError:
        print(f"  {RED}[!] Message class '{service_name}' not found{RESET}")
        return None

    data = {k: v for k, v in (auth_data or [])}
    for field_name in grpc_module.get_request_variable_names(module):
        if field_name not in data:
            vt = grpc_module.get_request_variable_type(module.DESCRIPTOR.fields_by_name[field_name])
            data[field_name] = grpc_module.generate_data(vt)

    try:
        request = getattr(stub, auth_rpc)
        resp    = await request(module(**data), metadata=None, timeout=timeout)
        result  = grpc_module.getDataResult(resp)
    except grpc.RpcError as e:
        print(f"  {RED}[!] Auth flow failed: {e.details()}{RESET}")
        return None
    except Exception as e:
        print(f"  {RED}[!] Auth flow failed: {e}{RESET}")
        return None

    token = result.get(auth_field)
    if not token:
        print(f"  {RED}[!] Field '{auth_field}' not found in response: {result}{RESET}")
        return None

    if auth_header:
        key, val_template = auth_header.split('=', 1)
        return [(key.strip().lower(), val_template.strip().replace('{token}', str(token)))]
    return [('authorization', f'Bearer {token}')]


# ── Baseline Measurement ───────────────────────────────────────────────────────

async def _measure_baselines(test_plan, metadata, timeout, samples=3):
    """Fix #8: measure baseline response time per service for accurate time-based detection."""
    baselines = {}
    seen = set()

    for tc in test_plan:
        service = tc['service']
        if service in seen:
            continue
        seen.add(service)

        has_time = any(t['time_threshold'] > 0 for t in test_plan if t['service'] == service)
        if not has_time:
            baselines[service] = 0.0
            continue

        module = tc['module']
        default_data = {
            key: grpc_module.generate_data(
                grpc_module.get_request_variable_type(module.DESCRIPTOR.fields_by_name[key])
            )
            for key in grpc_module.get_request_variable_names(module)
        }

        times = []
        stub = tc['stub']
        for _ in range(samples):
            t0 = time.time()
            try:
                req  = getattr(stub, service.split("Request")[0])
                await req(module(**default_data), metadata=metadata, timeout=timeout)
            except Exception:
                pass
            times.append((time.time() - t0) * 1000)

        baseline = sum(times) / len(times)
        baselines[service] = baseline
        print(f"  -> {service.replace('Request','')}: baseline {baseline:.0f} ms")

    return baselines


# ── Test Plan ──────────────────────────────────────────────────────────────────

def _build_test_plan(request_classes, gs_service, temp_payloads, stub_map,
                     attacks_filter=None, services_filter=None,
                     payloads_path='./core/modules/payloads.yaml'):
    """Phase 3: build flat list of test cases with filters applied."""
    test_cases    = []
    norm_attacks  = [a.lower() for a in attacks_filter]  if attacks_filter  else None
    norm_services = [s.lower() for s in services_filter] if services_filter else None

    for service in request_classes:
        if norm_services and service.replace("Request", "").lower() not in norm_services:
            continue

        stub = stub_map.get(service)
        if stub is None:
            print(f"  {DIM}[!] No stub for '{service}', skipping{RESET}")
            continue

        module       = getattr(gs_service, service)
        param_names  = grpc_module.get_request_variable_names(module)
        default_data = {
            key: grpc_module.generate_data(
                grpc_module.get_request_variable_type(module.DESCRIPTOR.fields_by_name[key])
            )
            for key in param_names
        }

        for item in param_names:
            var_type = grpc_module.get_request_variable_type(module.DESCRIPTOR.fields_by_name[item])
            if var_type == 'bool':
                continue
            if var_type is None:
                print(f"  {DIM}[!] Unsupported field type for '{item}' in {service}{RESET}")
                continue

            # Fix #10: specific exception handling instead of silent swallow
            try:
                attack_payloads = payload.getPayload(
                    varType=var_type, payloads=temp_payloads, payloads_path=payloads_path
                )
            except KeyError as e:
                print(f"  {DIM}[!] No mapping for type '{var_type}': {e}{RESET}")
                continue
            except Exception as e:
                print(f"  {DIM}[!] Payload load error for '{item}': {e}{RESET}")
                continue

            for attack, attack_data in attack_payloads.items():
                if norm_attacks and attack.lower() not in norm_attacks:
                    continue
                for payload_param in attack_data['param']:
                    combined = payload.combinePayload(default_data[item], payload_param)
                    # Fix #3: skip if combinePayload returns the unchanged default
                    # (means payload type is unsupported for this field)
                    if combined == default_data[item] and not isinstance(payload_param, (str, int, float)):
                        continue
                    test_data       = copy.copy(default_data)
                    test_data[item] = combined
                    test_cases.append({
                        'service':        service,
                        'stub':           stub,
                        'module':         module,
                        'param_names':    param_names,
                        'param':          item,
                        'attack':         attack,
                        'payload_param':  payload_param,
                        'test_data':      test_data,
                        'excepted_resp':  attack_data['resp'],
                        'exclude_resp':   attack_data.get('exclude', []),
                        'time_threshold': attack_data.get('time', 0),
                    })
    return test_cases


# ── Excel Export ───────────────────────────────────────────────────────────────

def generate_excel_result(results, meta):
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename  = f"vulnerability_list_{timestamp}.xlsx"

    for entry in results:
        if isinstance(entry['Response'], dict):
            entry['Response'] = str(entry['Response'])

    # Fix #9: mark duplicates before writing Excel
    seen_findings = {}
    for r in results:
        if r['Status_vulnerability'] != 'Vulnerable':
            r['Is_Duplicate'] = False
            continue
        key = (r['Service'], r['Vulnerable_param'], r['Attack_title'])
        if key in seen_findings:
            r['Is_Duplicate'] = True
        else:
            seen_findings[key] = r
            r['Is_Duplicate'] = False

    df = pd.DataFrame(results)

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # ── Sheet 1: Results ──────────────────────────────────────────────────
        df.to_excel(writer, sheet_name='Results', index=False)
        ws = writer.sheets['Results']

        hdr_fill = PatternFill(start_color="2F4F8F", end_color="2F4F8F", fill_type="solid")
        hdr_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal='center')

        red_fill    = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        green_fill  = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        orange_fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
        dim_fill    = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        status_idx  = list(df.columns).index('Status_vulnerability')
        conf_idx    = list(df.columns).index('Confidence') if 'Confidence' in df.columns else None
        dup_idx     = list(df.columns).index('Is_Duplicate') if 'Is_Duplicate' in df.columns else None

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            status = row[status_idx].value
            conf   = row[conf_idx].value if conf_idx is not None else 0
            is_dup = row[dup_idx].value  if dup_idx  is not None else False
            if status == "Vulnerable":
                fill = dim_fill if is_dup else (red_fill if conf >= 75 else orange_fill)
            else:
                fill = green_fill
            for cell in row:
                cell.fill = fill

        for col_idx, col in enumerate(df.columns, 1):
            max_len = max(df[col].astype(str).map(len).max(), len(col)) + 4
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 60)
        ws.freeze_panes = 'A2'

        # ── Sheet 2: Unique Findings ──────────────────────────────────────────
        unique_vulns = list(seen_findings.values())
        if unique_vulns:
            df_unique = pd.DataFrame(unique_vulns)
            df_unique = df_unique[df_unique['Status_vulnerability'] == 'Vulnerable'] \
                .drop(columns=['Is_Duplicate'], errors='ignore') \
                .sort_values('Confidence', ascending=False)
            df_unique.to_excel(writer, sheet_name='Unique Findings', index=False)
            ws_u = writer.sheets['Unique Findings']
            for cell in ws_u[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal='center')
            for col_idx, col in enumerate(df_unique.columns, 1):
                max_len = max(df_unique[col].astype(str).map(len).max(), len(col)) + 4
                ws_u.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 60)
            ws_u.freeze_panes = 'A2'

        # ── Sheet 3: Summary ──────────────────────────────────────────────────
        ws_s       = writer.book.create_sheet('Summary')
        vuln_count = sum(1 for r in results if r['Status_vulnerability'] == 'Vulnerable')
        not_vuln   = len(results) - vuln_count
        unique_cnt = len(seen_findings)
        high_conf  = sum(1 for r in results if r.get('Confidence', 0) >= 75
                         and r['Status_vulnerability'] == 'Vulnerable'
                         and not r.get('Is_Duplicate'))
        time_det   = sum(1 for r in results if 'time' in str(r.get('Detection_Method', ''))
                         and not r.get('Is_Duplicate'))

        summary_rows = [
            ("GSTF Scan Summary",    ""),
            ("",                     ""),
            ("Target",               meta['url']),
            ("Proto File",           meta['pathname']),
            ("Scan Date",            meta['timestamp']),
            ("Total Execution Time", f"{meta['total_time_ms']} ms"),
            ("",                     ""),
            ("Total Planned",          meta.get('total_planned', len(results))),
            ("  Executed",            len(results)),
            ("  Skipped (early exit)", meta.get('skipped', 0)),
            ("Vulnerable (total)",    vuln_count),
            ("Unique Vulnerabilities", unique_cnt),
            ("  High Confidence",     high_conf),
            ("  Time-Based",          time_det),
            ("Not Vulnerable",        not_vuln),
            ("Vulnerability Rate",    f"{vuln_count/len(results)*100:.1f}%" if results else "0%"),
            ("",                      ""),
            ("Services Tested",       ", ".join(s.replace("Request","") for s in meta['services'])),
            ("Attack Types",          ", ".join(meta['attack_types'])),
            ("",                      ""),
            ("Timeout per Request",   f"{meta.get('timeout', 30)} s"),
            ("Delay per Request",     f"{meta.get('delay_ms', 0)} ms"),
            ("Early Exit Threshold",  f"{meta.get('max_vuln_per_attack', 2)} vuln/attack"
                                      if meta.get('max_vuln_per_attack', 0) > 0 else "disabled"),
            ("Proxy",                 meta.get('proxy', "None")),
            ("Auth Flow",             meta.get('auth_rpc') or "None"),
        ]
        for r_idx, (label, value) in enumerate(summary_rows, 1):
            c1 = ws_s.cell(row=r_idx, column=1, value=label)
            ws_s.cell(row=r_idx, column=2, value=str(value) if value != "" else "")
            if label and value == "":
                c1.font = Font(bold=True, size=11)
        ws_s.column_dimensions['A'].width = 28
        ws_s.column_dimensions['B'].width = 50

    return filename, vuln_count, not_vuln, unique_cnt


# ── Request Execution ──────────────────────────────────────────────────────────

async def create_request(stub, module, service, data, excepted_resp, exclude_resp,
                         attack_title, vulnerable_param, payload_param, metadata,
                         index, total, quiet, time_threshold, baseline_ms, timeout):
    output   = None
    rpc_code = 0

    t0 = time.time()
    try:
        request = getattr(stub, service.split("Request")[0])
        # Fix #1: enforce per-request timeout
        resp   = await request(module(**data), metadata=metadata, timeout=timeout)
        output = grpc_module.getDataResult(resp)
    except grpc.RpcError as e:
        rpc_code = e.code()
        output   = str(e.details())
    except Exception as e:
        rpc_code = "UNKNOWN"
        output   = str(e)
    elapsed_ms = (time.time() - t0) * 1000

    output_str = str(output).lower()
    matches    = [item for item in excepted_resp if str(item).lower() in output_str]
    is_vulnerable = bool(matches)

    if is_vulnerable:
        payload_lower = str(payload_param).lower()
        if all(str(m).lower() in payload_lower for m in matches):
            is_vulnerable = False

    if is_vulnerable and exclude_resp:
        if any(str(ex).lower() in output_str for ex in exclude_resp):
            is_vulnerable = False

    # Fix #8: time-based uses baseline to reduce false positives under load
    effective_threshold_ms = max(
        time_threshold * 1000,
        baseline_ms + time_threshold * 1000
    ) if baseline_ms > 0 else time_threshold * 1000

    is_time_based = (
        time_threshold > 0 and
        elapsed_ms >= effective_threshold_ms and
        rpc_code == 0
    )
    if is_time_based:
        is_vulnerable = True

    confidence = _calculate_confidence(matches if is_vulnerable else [], is_time_based)
    if is_vulnerable:
        detection_method = (
            "string+time" if (is_time_based and matches) else
            "time_based"  if is_time_based else
            "string_match"
        )
        status = "Vulnerable"
    else:
        detection_method = "N/A"
        status = "Not Vulnerable"

    progress    = f"[{index:>4}/{total}]"
    atk_label   = attack_title.ljust(20)
    param_label = vulnerable_param.ljust(14)

    if is_vulnerable:
        conf_tag   = f"[{confidence}% {_confidence_label(confidence)}]"
        short_resp = str(output)
        short_resp = (short_resp[:60] + "...") if len(short_resp) > 60 else short_resp
        time_tag   = f" +time({elapsed_ms:.0f}ms)" if is_time_based else ""
        print(f"  {progress} {YELLOW}{atk_label}{RESET} | {param_label} | "
              f"{RED}VULNERABLE{RESET} {conf_tag}{time_tag} | {short_resp}")
    elif not quiet:
        print(f"  {DIM}{progress} {atk_label} | {param_label} | not vuln   |{RESET}")

    return {
        'Status_vulnerability': status,
        'Confidence':           confidence,
        'Detection_Method':     detection_method,
        'RPC_code':             rpc_code,
        'RPC_status':           "Error" if rpc_code != 0 else "Success",
        'Payload':              payload_param,
        'Response':             output,
        'Attack_title':         attack_title,
        'Service':              service,
        'Vulnerable_param':     vulnerable_param,
        'Elapsed_ms':           int(elapsed_ms),
    }


# ── Stub Builder ───────────────────────────────────────────────────────────────

def generate_stub(url, pathname, secure, proxy=None):
    """Fix #2: build stub_map supporting multiple services in one proto file."""
    module_name = grpc_module.loadProto(pathname)
    base_name   = module_name.split('.')[0]

    gs_service = grpc_module.import_proto_module(f'{base_name}_pb2')
    gm_service = grpc_module.import_proto_module(f'{base_name}_pb2_grpc')

    request_classes = grpc_module.get_request_class_name(f'{base_name}_pb2')
    stub_names      = grpc_module.get_all_class_stubs(f'{base_name}_pb2_grpc')

    all_stubs = [
        grpc_module.getStub(url=url, message_stub=getattr(gm_service, sn),
                            secure=secure, proxy=proxy)
        for sn in stub_names
    ]

    stub_map = {}
    for rc in request_classes:
        method = rc.split("Request")[0]
        for stub in all_stubs:
            if hasattr(stub, method):
                stub_map[rc] = stub
                break

    return stub_map, request_classes, gs_service


# ── Main Scan ──────────────────────────────────────────────────────────────────

async def start_scan(
    pathname,
    secure,
    url,
    metadata             = None,
    quiet                = False,
    delay_ms             = 0,
    attacks_filter       = None,
    services_filter      = None,
    payloads_path        = './core/modules/payloads.yaml',
    auth_rpc             = None,
    auth_data            = None,
    auth_field           = 'token',
    auth_header          = None,
    timeout              = 30,
    max_vuln_per_attack  = 2,
    proxy                = None,
):
    scan_start  = time.time()
    results     = []
    all_attacks = set()

    if not metadata:
        metadata = None

    _print_banner(url, pathname, metadata,
                  {'attacks': attacks_filter, 'services': services_filter},
                  proxy=proxy)

    # Phase 1: Extraction of Payload
    print(f"{BOLD}[PHASE 1]{RESET} Extracting payloads...")
    temp_payloads = payload.loadPayload(payloads_path)
    attack_count  = len(temp_payloads['payload'])
    payload_count = sum(len(v['param']) for v in temp_payloads['payload'].values())
    print(f"  -> {attack_count} attack types | {payload_count} payloads total\n")

    # Phase 2: Code Generation
    print(f"{BOLD}[PHASE 2]{RESET} Compiling proto & generating stub...")
    t0 = time.time()
    stub_map, request_classes, gs_service = generate_stub(url, pathname, secure, proxy=proxy)
    init_ms    = int((time.time() - t0) * 1000)
    svc_labels = ", ".join(r.replace("Request", "") for r in request_classes)
    print(f"  -> Done in {init_ms} ms | Services: {svc_labels}\n")

    # Auth Flow (optional)
    if auth_rpc:
        print(f"{BOLD}[AUTH]  {RESET} Running authentication flow ({auth_rpc})...")
        acquired = await _run_auth_flow(
            stub_map, gs_service, auth_rpc, auth_data, auth_field, auth_header, timeout
        )
        if acquired:
            metadata = (list(metadata) if metadata else []) + acquired
            print(f"  -> Token acquired | {_mask_metadata(acquired)}\n")
        else:
            print(f"  {YELLOW}[!] Auth flow failed — continuing without token{RESET}\n")

    # Phase 3: Test Case Creation
    print(f"{BOLD}[PHASE 3]{RESET} Building test cases...")
    test_plan = _build_test_plan(
        request_classes, gs_service, temp_payloads, stub_map,
        attacks_filter=attacks_filter, services_filter=services_filter,
        payloads_path=payloads_path,
    )
    total = len(test_plan)
    filter_note = ""
    if attacks_filter:
        filter_note += f" | attacks: {', '.join(attacks_filter)}"
    if services_filter:
        filter_note += f" | services: {', '.join(services_filter)}"
    print(f"  -> {total} test cases across {len(request_classes)} services{filter_note}\n")

    if delay_ms > 0:
        print(f"  {DIM}[i] Delay: {delay_ms}ms/req "
              f"(est. +{total * delay_ms / 1000:.0f}s){RESET}\n")

    # Baseline measurement for time-based detection (Fix #8)
    has_time_attacks = any(tc['time_threshold'] > 0 for tc in test_plan)
    baselines = {}
    if has_time_attacks:
        print(f"{BOLD}[BASELINE]{RESET} Measuring response baselines for time-based detection...")
        baselines = await _measure_baselines(test_plan, metadata, timeout)
        print()

    # Phase 4: Execution
    early_exit_note = (f" | early exit after {max_vuln_per_attack} vuln/attack"
                       if max_vuln_per_attack > 0 else "")
    print(f"{BOLD}[PHASE 4]{RESET} Executing tests...  "
          f"[timeout: {timeout}s/req{early_exit_note}]")
    print("-" * 64)

    current_service = None
    svc_start       = time.time()
    svc_vuln        = 0
    vuln_per_group  = {}   # (service, param, attack) -> count of vulnerable payloads found
    skip_notified   = set()
    skipped_count   = 0

    for index, tc in enumerate(test_plan, 1):
        service = tc['service']

        # Always show service header before any early-exit check
        if service != current_service:
            if current_service is not None:
                svc_ms = int((time.time() - svc_start) * 1000)
                vc     = RED if svc_vuln > 0 else GREEN
                print(f"\n  -> {current_service.replace('Request','')}: "
                      f"{svc_ms} ms | {vc}{svc_vuln} vulnerabilities found{RESET}")
            current_service = service
            svc_start       = time.time()
            svc_vuln        = 0
            print(f"\n  {BOLD}>> {service.replace('Request','')}{RESET}"
                  f"  (params: {', '.join(tc['param_names'])})")

        # Early exit: skip remaining payloads once threshold is reached for this group
        group_key = (service, tc['param'], tc['attack'])
        if max_vuln_per_attack > 0 and vuln_per_group.get(group_key, 0) >= max_vuln_per_attack:
            skipped_count += 1
            if group_key not in skip_notified:
                skip_notified.add(group_key)
                remaining = sum(
                    1 for t in test_plan[index - 1:]
                    if (t['service'], t['param'], t['attack']) == group_key
                )
                if not quiet:
                    print(f"  {DIM}[->] Early exit: {tc['attack'].upper()} on "
                          f"'{tc['param']}' — {max_vuln_per_attack} found, "
                          f"skipping {remaining} remaining payload(s){RESET}")
            continue

        result = await create_request(
            stub=tc['stub'], module=tc['module'], service=service,
            data=tc['test_data'], excepted_resp=tc['excepted_resp'],
            exclude_resp=tc['exclude_resp'], attack_title=tc['attack'].upper(),
            vulnerable_param=tc['param'], payload_param=tc['payload_param'],
            metadata=metadata, index=index, total=total,
            quiet=quiet, time_threshold=tc['time_threshold'],
            baseline_ms=baselines.get(service, 0.0), timeout=timeout,
        )
        result['No'] = index
        results.append(result)
        all_attacks.add(tc['attack'])
        if result['Status_vulnerability'] == 'Vulnerable':
            svc_vuln += 1
            vuln_per_group[group_key] = vuln_per_group.get(group_key, 0) + 1

        if delay_ms > 0:
            await asyncio.sleep(delay_ms / 1000)

    if current_service is not None:
        svc_ms = int((time.time() - svc_start) * 1000)
        vc     = RED if svc_vuln > 0 else GREEN
        print(f"\n  -> {current_service.replace('Request','')}: "
              f"{svc_ms} ms | {vc}{svc_vuln} vulnerabilities found{RESET}")

    total_ms = int((time.time() - scan_start) * 1000)

    seen = set()
    ordered_services = [
        tc['service'] for tc in test_plan
        if not (tc['service'] in seen or seen.add(tc['service']))
    ]

    executed_count = len(results)
    meta = {
        'url':                url,
        'pathname':           pathname,
        'timestamp':          datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'total_time_ms':      total_ms,
        'services':           ordered_services,
        'attack_types':       sorted(all_attacks),
        'delay_ms':           delay_ms,
        'auth_rpc':           auth_rpc,
        'timeout':            timeout,
        'total_planned':      total,
        'skipped':            skipped_count,
        'max_vuln_per_attack': max_vuln_per_attack,
        'proxy':              proxy or "None",
    }

    filename, vuln_count, not_vuln_count, unique_count = generate_excel_result(results, meta)
    vuln_rate  = vuln_count / executed_count * 100 if executed_count else 0
    high_conf  = sum(1 for r in results if r.get('Confidence', 0) >= 75
                     and r['Status_vulnerability'] == 'Vulnerable'
                     and not r.get('Is_Duplicate'))
    time_based = sum(1 for r in results if 'time' in str(r.get('Detection_Method', ''))
                     and not r.get('Is_Duplicate'))

    print(f"\n{BOLD}{CYAN}{SEP}")
    print(f"  SCAN COMPLETE")
    print(f"{SEP}{RESET}")
    print(f"  {'Total Planned':<30}: {total}")
    print(f"  {'  Executed':<30}: {executed_count}")
    if skipped_count:
        print(f"  {'  Skipped (early exit)':<30}: {DIM}{skipped_count}{RESET}")
    print(f"  {'Vulnerable (total)':<30}: {RED}{vuln_count} ({vuln_rate:.1f}%){RESET}")
    print(f"  {'Unique Vulnerabilities':<30}: {RED}{unique_count}{RESET}")
    if unique_count:
        print(f"  {'  High Confidence (>=75%)':<30}: {RED}{high_conf}{RESET}")
        if time_based:
            print(f"  {'  Time-Based Detected':<30}: {RED}{time_based}{RESET}")
    print(f"  {'Not Vulnerable':<30}: {GREEN}{not_vuln_count}{RESET}")
    print(f"  {'Total Time':<30}: {total_ms} ms")
    print(f"  {'Report':<30}: {BOLD}{filename}{RESET}")
    print(f"{BOLD}{CYAN}{SEP}{RESET}\n")

    return {
        'total':      total,
        'executed':   executed_count,
        'skipped':    skipped_count,
        'vulnerable': vuln_count,
        'unique':     unique_count,
        'report':     filename,
    }
